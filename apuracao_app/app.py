"""App de Apuração de Sorteio.

Cruza as dezenas sorteadas de cada prêmio (Ata de Sorteio) com as cartelas
comercializadas (arquivo de Comercializados) e identifica as cartelas contempladas,
prêmio a prêmio, com validação cruzada contra o gabarito da própria Ata de Sorteio.
"""
from __future__ import annotations

import io
import json
import os
import re
from datetime import datetime

import pandas as pd
import streamlit as st

from apuracao import apurar_extracao, validar_extracao
from parsers import ParseError, extrair_edicao, parse_comercializados, parse_sorteio

APP_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_FOLDER = os.path.dirname(APP_DIR)
# Config no perfil do usuário: o app pode estar instalado em pasta sem permissão de
# escrita (ex: C:\Program Files), então nada de estado gravado junto do código.
CONFIG_FILE = os.path.join(os.path.expanduser("~"), ".apuracao_sorteio.json")


def load_last_folder() -> str:
    try:
        with open(CONFIG_FILE, encoding="utf-8") as f:
            folder = json.load(f).get("folder")
            if folder and os.path.isdir(folder):
                return folder
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        pass
    return DEFAULT_FOLDER


def save_last_folder(folder: str) -> None:
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump({"folder": folder}, f)
    except OSError:
        pass


def _walk_limited(root: str, max_depth: int = 4):
    root = os.path.abspath(root)
    base_depth = root.rstrip(os.sep).count(os.sep)
    for dirpath, dirnames, filenames in os.walk(root):
        depth = dirpath.rstrip(os.sep).count(os.sep) - base_depth
        if depth >= max_depth:
            dirnames[:] = []
        for filename in filenames:
            yield os.path.join(dirpath, filename)


def find_files(folder: str, keyword: str):
    if not folder or not os.path.isdir(folder):
        return []
    return sorted(
        p for p in _walk_limited(folder)
        if p.lower().endswith(".txt") and keyword in os.path.basename(p).lower()
    )


def _fmt_data(raw: str) -> str:
    if raw and len(raw) == 8 and raw.isdigit():
        return f"{raw[0:2]}/{raw[2:4]}/{raw[4:8]}"
    return raw or "-"


def _ordinal(codigo: str) -> str:
    numero = int(codigo[1:])
    return f"{numero}º Prêmio"


def _fmt_dezenas(dezenas) -> str:
    return ", ".join(f"{d:02d}" for d in sorted(dezenas))


def _nome_arquivo(source) -> str:
    return getattr(source, "name", None) or (source if isinstance(source, str) else "") or ""


@st.cache_data(show_spinner=False)
def _comercializados_do_caminho(caminho: str, mtime: float, tamanho: int, _esquema: int = 2):
    """Cache da leitura do arquivo grande, invalidado quando o arquivo muda no disco
    (mtime/tamanho entram na chave do cache). Só usado no modo 'Apontar pasta'.

    `_esquema`: incrementar sempre que o formato retornado por parse_comercializados
    mudar — o st.cache_data só invalida pelo código DESTA função, não das chamadas."""
    return parse_comercializados(caminho)


def _carregar_comercializados(source):
    if isinstance(source, str):
        stat = os.stat(source)
        return _comercializados_do_caminho(source, stat.st_mtime, stat.st_size)
    return parse_comercializados(source)


def _fmt_milhar(n: int) -> str:
    return f"{n:,}".replace(",", ".")


def _versao_layout_exibicao(header: dict) -> str:
    versao = header.get("versao_layout") or ""
    if versao not in ("01", "02"):
        versao = f"{header.get('layout_detectado', '-')} (auto)"
    return versao


def _gerar_excel(resultados: dict, header: dict, edicao: str,
                 total_titulos: int, total_cartelas: int, avisos: list) -> io.BytesIO:
    """Gera o relatório da apuração em uma única aba formatada (estilo relatório),
    com a mesma paleta do app: título, bloco de informações (cards), avisos de
    integridade e uma seção por extração com a tabela de certificados contemplados."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    AZUL, AZUL_ESCURO, AZUL_CLARO = "2563EB", "1E3A8A", "EFF6FF"
    VERDE, VERDE_TXT = "DCFCE7", "166534"
    AMBAR, AMBAR_TXT = "FEF3C7", "92400E"
    CINZA, BRANCO = "64748B", "FFFFFF"

    def fundo(cor):
        return PatternFill("solid", fgColor=cor)

    lado = Side(style="thin", color="93C5FD")
    borda = Border(left=lado, right=lado, top=lado, bottom=lado)
    wrap = Alignment(wrap_text=True, vertical="center")
    centro = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "Apuração"
    ws.sheet_view.showGridLines = False
    for col, largura in {"A": 2.5, "B": 6, "C": 18, "D": 18, "E": 26, "F": 34}.items():
        ws.column_dimensions[col].width = largura

    def linha_mesclada(row, texto, *, col_ini=2, col_fim=6, fonte=None, cor_fundo=None,
                       altura=None, alinhamento=None, com_borda=False):
        for c in range(col_ini, col_fim + 1):
            cel = ws.cell(row=row, column=c)
            if cor_fundo:
                cel.fill = fundo(cor_fundo)
            if fonte:
                cel.font = fonte
            if com_borda:
                cel.border = borda
        cel = ws.cell(row=row, column=col_ini, value=texto)
        if alinhamento:
            cel.alignment = alinhamento
        if altura:
            ws.row_dimensions[row].height = altura
        if col_fim > col_ini:
            ws.merge_cells(start_row=row, start_column=col_ini, end_row=row, end_column=col_fim)
        return cel

    # ---- Título ----
    linha = 2
    linha_mesclada(linha, "Apuração de Sorteio",
                   fonte=Font(bold=True, size=16, color=BRANCO), cor_fundo=AZUL,
                   altura=30, alinhamento=Alignment(vertical="center"))
    linha += 1
    gerado_em = datetime.now().strftime("%d/%m/%Y %H:%M")
    linha_mesclada(linha, f"Edição {edicao}  •  relatório gerado em {gerado_em}",
                   fonte=Font(size=10, color="DBEAFE"), cor_fundo=AZUL,
                   altura=18, alinhamento=Alignment(vertical="center"))
    linha += 2

    # ---- Bloco de informações (cards da tela) ----
    infos = [
        ("Edição", edicao or "-"),
        ("Empresa", header.get("nome_empresa") or "-"),
        ("Data do sorteio", _fmt_data(header.get("data_sorteio"))),
        ("Versão do layout", _versao_layout_exibicao(header)),
        ("Títulos comercializados", _fmt_milhar(total_titulos)),
        ("Cartelas comercializadas lidas", _fmt_milhar(total_cartelas)),
    ]
    for rotulo, valor in infos:
        for c in range(2, 7):
            cel = ws.cell(row=linha, column=c)
            cel.fill = fundo(AZUL_CLARO)
            cel.border = borda
        cel_rotulo = ws.cell(row=linha, column=2, value=rotulo)
        cel_rotulo.font = Font(bold=True, size=10, color=AZUL_ESCURO)
        ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=3)
        cel_valor = ws.cell(row=linha, column=4, value=str(valor))
        cel_valor.font = Font(size=10)
        ws.merge_cells(start_row=linha, start_column=4, end_row=linha, end_column=6)
        ws.row_dimensions[linha].height = 17
        linha += 1

    # ---- Avisos de integridade ----
    if avisos:
        linha += 1
        for aviso in avisos:
            linha_mesclada(linha, f"⚠ {aviso}", fonte=Font(size=9, color=AMBAR_TXT),
                           cor_fundo=AMBAR, altura=28, alinhamento=wrap, com_borda=True)
            linha += 1

    # ---- Seções por extração ----
    for codigo, dados in resultados.items():
        extracao = dados["extracao"]
        vencedoras = dados["vencedoras"]
        validacao = dados["validacao"]
        linha += 1
        linha_mesclada(linha, f"{codigo} — {_ordinal(codigo)}",
                       fonte=Font(bold=True, size=12, color=BRANCO), cor_fundo=AZUL_ESCURO,
                       altura=22, alinhamento=Alignment(vertical="center"))
        linha += 1
        linha_mesclada(
            linha,
            f"Propostas premiadas (arquivo): {extracao.propostas_premiadas}   •   "
            f"Cartelas vencedoras (calculado): {len(vencedoras)}",
            fonte=Font(size=10, color=CINZA), altura=16,
        )
        linha += 1
        linha_mesclada(
            linha,
            f"Dezenas sorteadas ({len(extracao.dezenas)}): {_fmt_dezenas(extracao.dezenas)}",
            fonte=Font(size=9, color=CINZA), altura=28, alinhamento=wrap,
        )
        linha += 1
        if validacao["ok"]:
            linha_mesclada(
                linha,
                f"✓ Validação OK — confere com o gabarito da Ata de Sorteio "
                f"({validacao['qtd_esperada']} certificado(s)).",
                fonte=Font(size=10, bold=True, color=VERDE_TXT), cor_fundo=VERDE,
                altura=18, alinhamento=wrap, com_borda=True,
            )
        else:
            detalhes = []
            if validacao["cert_faltando"]:
                detalhes.append("certificados da Ata não encontrados: " + ", ".join(validacao["cert_faltando"]))
            if validacao["cert_extra"]:
                detalhes.append("apurados ausentes na Ata: " + ", ".join(validacao["cert_extra"]))
            linha_mesclada(
                linha,
                "⚠ Divergência em relação à Ata de Sorteio — " + "; ".join(detalhes),
                fonte=Font(size=10, bold=True, color=AMBAR_TXT), cor_fundo=AMBAR,
                altura=30, alinhamento=wrap, com_borda=True,
            )
        linha += 2

        if vencedoras:
            for c, titulo in ((2, "Nº"), (3, "Certificado"), (4, "Número da sorte")):
                cel = ws.cell(row=linha, column=c, value=titulo)
                cel.font = Font(bold=True, size=10, color=BRANCO)
                cel.fill = fundo(AZUL)
                cel.border = borda
                cel.alignment = centro
            for c in range(5, 7):
                cel = ws.cell(row=linha, column=c)
                cel.fill = fundo(AZUL)
                cel.border = borda
            cel = ws.cell(row=linha, column=5, value="Dezenas da cartela")
            cel.font = Font(bold=True, size=10, color=BRANCO)
            cel.alignment = centro
            ws.merge_cells(start_row=linha, start_column=5, end_row=linha, end_column=6)
            ws.row_dimensions[linha].height = 17
            linha += 1
            for n, v in enumerate(vencedoras, start=1):
                for c in range(2, 7):
                    ws.cell(row=linha, column=c).border = borda
                cel = ws.cell(row=linha, column=2, value=n)
                cel.alignment = centro
                cel.font = Font(size=10)
                cel = ws.cell(row=linha, column=3, value=v.certificado)
                cel.alignment = centro
                cel.font = Font(size=10, bold=True)
                cel = ws.cell(row=linha, column=4, value=v.numero_sorte)
                cel.alignment = centro
                cel.font = Font(size=10)
                cel = ws.cell(row=linha, column=5, value=_fmt_dezenas(v.dezenas))
                cel.font = Font(size=10)
                cel.alignment = wrap
                ws.merge_cells(start_row=linha, start_column=5, end_row=linha, end_column=6)
                ws.row_dimensions[linha].height = 16
                linha += 1
        else:
            linha_mesclada(linha, "Nenhuma cartela vencedora nesta extração.",
                           fonte=Font(size=10, italic=True, color=CINZA), altura=16)
            linha += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


st.set_page_config(page_title="Apuração de Sorteio", page_icon="🎯", layout="wide")

st.markdown(
    """
    <style>
        .stButton > button { border-radius: 8px; font-weight: 600; }
        div[data-testid="stMetric"] {
            background-color: #EFF6FF;
            border-radius: 10px;
            padding: 12px 16px;
        }
        h3 { color: #1E3A8A; }
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("🎯 Apuração de Sorteio")
st.caption(
    "Cruza as dezenas sorteadas de cada prêmio (Ata de Sorteio) com as cartelas comercializadas "
    "(Comercializados) e identifica as cartelas contempladas — extração por extração."
)

st.subheader("1. Arquivos de entrada")
modo = st.radio("Como deseja indicar os arquivos?", ["Apontar pasta", "Upload manual"], horizontal=True)

sorteio_source = None
comerc_source = None

if modo == "Apontar pasta":
    folder = st.text_input("Pasta onde estão os arquivos", value=load_last_folder())
    if folder and os.path.isdir(folder):
        save_last_folder(folder)
        sorteio_files = find_files(folder, "sorteio")
        comerc_files = find_files(folder, "comercializad")
        col1, col2 = st.columns(2)
        with col1:
            if sorteio_files:
                sorteio_source = st.selectbox(
                    "Arquivo Ata de Sorteio",
                    sorteio_files,
                    format_func=lambda p: os.path.relpath(p, folder),
                )
            else:
                st.warning("Nenhum arquivo com 'sorteio' no nome (.txt) encontrado nessa pasta.")
        with col2:
            if comerc_files:
                comerc_source = st.selectbox(
                    "Arquivo de Comercializados",
                    comerc_files,
                    format_func=lambda p: os.path.relpath(p, folder),
                )
            else:
                st.warning("Nenhum arquivo com 'comercializad' no nome (.txt) encontrado nessa pasta.")
    elif folder:
        st.error("Pasta não encontrada.")
else:
    col1, col2 = st.columns(2)
    with col1:
        sorteio_source = st.file_uploader("Arquivo Ata de Sorteio (.txt)", type=["txt"])
    with col2:
        comerc_source = st.file_uploader("Arquivo de Comercializados (.txt)", type=["txt"])

fonte_atual = (_nome_arquivo(sorteio_source), _nome_arquivo(comerc_source))

processar = st.button("▶️ Processar apuração", type="primary", disabled=not (sorteio_source and comerc_source))

if processar:
    try:
        with st.spinner("Lendo Ata de Sorteio..."):
            sorteio_data = parse_sorteio(sorteio_source)
        if not sorteio_data.extracoes:
            st.error(
                "Nenhuma extração de prêmio principal foi encontrada na Ata de Sorteio. "
                "Confira se o arquivo selecionado é realmente uma Ata de Sorteio no formato esperado."
            )
            st.stop()
        with st.spinner("Lendo arquivo de Comercializados (arquivo grande, pode levar alguns segundos)..."):
            comerc = _carregar_comercializados(comerc_source)
        if not comerc.cartelas:
            st.error(
                "Nenhuma cartela foi encontrada no arquivo de Comercializados. "
                "Confira se o arquivo selecionado é realmente um arquivo de Comercializados."
            )
            st.stop()
    except ParseError as e:
        st.error(f"Não foi possível processar: {e}")
        st.stop()
    except OSError as e:
        st.error(f"Não foi possível ler o arquivo: {e}")
        st.stop()

    resultados = {}
    for codigo in sorteio_data.codigos_extracoes():
        extracao = sorteio_data.extracoes[codigo]
        vencedoras = apurar_extracao(extracao.dezenas, comerc.cartelas)
        gabarito = sorteio_data.gabarito.get(codigo, [])
        validacao = validar_extracao(vencedoras, gabarito)
        resultados[codigo] = {"extracao": extracao, "vencedoras": vencedoras, "validacao": validacao}

    edicao = extrair_edicao(_nome_arquivo(sorteio_source)) or extrair_edicao(_nome_arquivo(comerc_source))
    if not edicao:
        # nome do arquivo sem série (Nomenclatura 1 do manual) — mostra a data do sorteio
        edicao = _fmt_data(sorteio_data.header.get("data_sorteio"))

    avisos = list(sorteio_data.avisos) + list(comerc.avisos)
    layout_sorteio = sorteio_data.header.get("layout_detectado")
    if comerc.versao_layout in ("01", "02") and comerc.versao_layout != layout_sorteio:
        avisos.append(
            f"A Ata de Sorteio usa o layout {layout_sorteio}, mas o arquivo de Comercializados declara versão "
            f"{comerc.versao_layout} — confira se os dois arquivos são da mesma edição."
        )

    st.session_state["sorteio_header"] = sorteio_data.header
    st.session_state["resultados"] = resultados
    st.session_state["total_cartelas"] = len(comerc.cartelas)
    st.session_state["total_titulos"] = comerc.total_titulos
    st.session_state["edicao"] = edicao
    st.session_state["avisos"] = avisos
    st.session_state["fonte_resultados"] = fonte_atual

resultados_atuais = (
    "resultados" in st.session_state and st.session_state.get("fonte_resultados") == fonte_atual
)

if "resultados" in st.session_state and not resultados_atuais:
    st.info("A seleção de arquivos mudou — clique em \"Processar apuração\" para atualizar os resultados.")

if resultados_atuais:
    st.subheader("2. Resultado da apuração")

    for aviso in st.session_state.get("avisos") or []:
        st.warning(f"⚠️ {aviso}")

    header = st.session_state.get("sorteio_header") or {}
    if header:
        with st.container(border=True):
            c1, c2, c3 = st.columns(3)
            c1.metric("Edição", st.session_state.get("edicao") or "-")
            c2.metric("Empresa", header.get("nome_empresa") or "-")
            c3.metric("Data do sorteio", _fmt_data(header.get("data_sorteio")))
            c4, c5, c6 = st.columns(3)
            c4.metric("Versão do layout", _versao_layout_exibicao(header))
            c5.metric("Títulos comercializados", _fmt_milhar(st.session_state.get("total_titulos", 0)))
            c6.metric("Cartelas comercializadas lidas", _fmt_milhar(st.session_state.get("total_cartelas", 0)))

    for codigo, dados in st.session_state["resultados"].items():
        extracao = dados["extracao"]
        vencedoras = dados["vencedoras"]
        validacao = dados["validacao"]
        with st.container(border=True):
            st.markdown(f"### 🏆 {codigo} — {_ordinal(codigo)}")
            c1, c2, c3 = st.columns(3)
            c1.metric("Dezenas sorteadas", len(extracao.dezenas))
            c2.metric("Propostas premiadas (arquivo)", extracao.propostas_premiadas)
            c3.metric("Cartelas vencedoras (calculado)", len(vencedoras))
            st.caption("Dezenas sorteadas: " + _fmt_dezenas(extracao.dezenas))

            if vencedoras:
                df = pd.DataFrame(
                    [{"Certificado": v.certificado, "Número da sorte": v.numero_sorte,
                      "Dezenas": _fmt_dezenas(v.dezenas)} for v in vencedoras]
                )
                st.dataframe(df, width="stretch", hide_index=True)
            else:
                st.info("Nenhuma cartela vencedora encontrada para esta extração.")

            if validacao["ok"]:
                st.success(
                    f"✅ Validação: os certificados apurados batem com a Ata de Sorteio "
                    f"({validacao['qtd_esperada']} certificado(s))."
                )
            else:
                st.warning("⚠️ Divergência em relação à Ata de Sorteio.")
                if validacao["cert_faltando"]:
                    st.write(
                        "Certificados da Ata **não encontrados** pelo programa: "
                        + ", ".join(validacao["cert_faltando"])
                    )
                if validacao["cert_extra"]:
                    st.write(
                        "Certificados apurados e **ausentes na Ata**: "
                        + ", ".join(validacao["cert_extra"])
                    )

    st.subheader("3. Exportar")
    excel_bytes = _gerar_excel(
        st.session_state["resultados"],
        st.session_state.get("sorteio_header") or {},
        st.session_state.get("edicao") or "-",
        st.session_state.get("total_titulos", 0),
        st.session_state.get("total_cartelas", 0),
        st.session_state.get("avisos") or [],
    )
    sufixo = re.sub(r"[^A-Za-z0-9._-]+", "_", st.session_state.get("edicao") or "sorteio")
    st.download_button(
        "⬇️ Baixar relatório em Excel",
        data=excel_bytes,
        file_name=f"apuracao_{sufixo}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
